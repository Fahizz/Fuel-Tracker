from __future__ import annotations

import json
from dataclasses import asdict
from datetime import date
from pathlib import Path
from typing import Union

from engine.models import PlanResult

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _serialize(obj):
    if isinstance(obj, date):
        return obj.isoformat()
    raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


def export_json(result: PlanResult, path: Union[str, Path]) -> None:
    data = asdict(result)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=_serialize)


def result_to_dict(result: PlanResult) -> dict:
    return json.loads(json.dumps(asdict(result), default=_serialize))


def export_excel(result: PlanResult, path: Union[str, Path]) -> None:
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export. pip install openpyxl")

    wb = openpyxl.Workbook()

    # --- Cycle Summary sheet ---
    ws = wb.active
    ws.title = "Cycle Summary"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = "Cycle Summary"
    ws["A1"].font = Font(bold=True, size=14)

    s = result.summary
    rows = [
        ("Cycle Start", str(s.cycle_start)),
        ("Cycle End", str(s.cycle_end)),
        ("Working Days", s.working_days),
        ("Office Commute KM", s.office_km),
        ("Extra Weekend KM", s.extra_weekend_km),
        ("Total KM Driven", s.total_km),
        ("Mileage (km/l)", s.mileage_kmpl),
        ("Fuel Used (litres)", s.fuel_used_litres),
        ("Raw Fuel Cost (INR)", s.raw_fuel_cost),
        ("Final Billed Total (INR)", s.final_billed_total),
        ("Start Odometer", s.start_odometer),
        ("End Odometer", s.end_odometer),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2, value=value)
        cell_a.font = Font(bold=True)
        cell_a.border = thin_border
        cell_b.border = thin_border

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    # --- Refuelling Breakdown sheet ---
    ws2 = wb.create_sheet("Refuelling Breakdown")
    headers = ["Refuel #", "Litres", "Price/Litre (INR)", "Amount (INR)", "Discount (INR)", "Final Bill (INR)"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(result.refuels, start=2):
        vals = [r.number, r.litres, r.price_per_litre, r.amount, r.discount, r.final_bill]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Totals row
    total_row = len(result.refuels) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=result.total_litres)
    ws2.cell(row=total_row, column=4, value=result.total_raw_cost)
    ws2.cell(row=total_row, column=5, value=round(sum(r.discount for r in result.refuels), 2))
    ws2.cell(row=total_row, column=6, value=result.total_final_billed)
    for col in range(1, 7):
        cell = ws2.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col_letter].width = 18

    wb.save(path)
